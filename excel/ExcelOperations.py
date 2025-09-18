"""
读取Excel总行数 - get_total_rows()
拷贝Excel内容 - copy_excel_content()
粘贴Excel内容 - paste_excel_content()
获取第一个可用行 - get_first_available_row()
获取上一个可用行 - get_first_available_row()（可调整逻辑）
获取第一个可用列 - get_first_available_col()
获取上一个可用列 - get_first_available_col()（可调整逻辑）
获取最后一列 - get_last_used_col()
清空Excel内容 - clear_excel_content()
删除Excel行 - delete_excel_row()
删除重复行 - delete_duplicate_rows()
删除所有行 - delete_all_rows()
删除Excel列 - delete_excel_column()
插入空行 - insert_empty_row()
插入空列 - insert_empty_column()
选中区域 - select_area()
获取选中区域 - get_selected_area()
设置行列隐藏 - set_column_hidden() 和 set_row_hidden()
"""

import openpyxl
from typing import List, Tuple, Optional
import pandas as pd


class ExcelOperations:
    """
    Excel操作类，提供一系列常用的Excel处理功能
    """

    def __init__(self, file_path: str = None):
        """
        初始化Excel操作类

        Args:
            file_path (str): Excel文件路径
        """
        self.file_path = file_path
        self.workbook = None
        self.active_sheet = None

    def open_file(self, file_path: str = None) -> bool:
        """
        打开Excel文件

        Args:
            file_path (str): 文件路径

        Returns:
            bool: 是否成功打开文件
        """
        try:
            if file_path:
                self.file_path = file_path
            self.workbook = openpyxl.load_workbook(self.file_path)
            self.active_sheet = self.workbook.active
            return True
        except Exception as e:
            print(f"打开文件失败: {e}")
            return False

    def close_file(self):
        """关闭Excel文件"""
        if self.workbook:
            self.workbook.save(self.file_path)
            self.workbook.close()

    def get_total_rows(self) -> int:
        """
        获取Excel总行数

        Returns:
            int: 总行数
        """
        if not self.active_sheet:
            return 0

        # 获取最后一行的行号
        last_row = self.active_sheet.max_row
        return last_row

    def get_total_cols(self) -> int:
        """
        获取Excel总列数

        Returns:
            int: 总列数
        """
        if not self.active_sheet:
            return 0

        # 获取最后一列的列号
        last_col = self.active_sheet.max_column
        return last_col

    def copy_excel_content(self, start_row: int = 1, start_col: int = 1,
                           end_row: int = None, end_col: int = None) -> List[List]:
        """
        拷贝Excel内容

        Args:
            start_row (int): 起始行
            start_col (int): 起始列
            end_row (int): 结束行
            end_col (int): 结束列

        Returns:
            List[List]: 拷贝的内容
        """
        if not self.active_sheet:
            return []

        # 设置默认结束位置
        if end_row is None:
            end_row = self.get_total_rows()
        if end_col is None:
            end_col = self.get_total_cols()

        content = []
        for row in range(start_row, end_row + 1):
            row_data = []
            for col in range(start_col, end_col + 1):
                cell_value = self.active_sheet.cell(row=row, column=col).value
                row_data.append(cell_value)
            content.append(row_data)

        return content

    def paste_excel_content(self, data: List[List], start_row: int = 1, start_col: int = 1):
        """
        粘贴Excel内容

        Args:
            data (List[List]): 要粘贴的数据
            start_row (int): 起始行
            start_col (int): 起始列
        """
        if not self.active_sheet or not data:
            return

        for i, row_data in enumerate(data):
            for j, cell_value in enumerate(row_data):
                self.active_sheet.cell(
                    row=start_row + i,
                    column=start_col + j
                ).value = cell_value

    def get_first_available_row(self, col_index: int = 1) -> int:
        """
        获取列中第一个可用行

        Args:
            col_index (int): 列索引

        Returns:
            int: 第一个可用行的行号
        """
        if not self.active_sheet:
            return 1

        for row in range(1, self.get_total_rows() + 1):
            cell_value = self.active_sheet.cell(row=row, column=col_index).value
            if cell_value is not None and str(cell_value).strip():
                return row

        return 1  # 如果没有找到，返回第一行

    def get_first_available_col(self, row_index: int = 1) -> int:
        """
        获取行中第一个可用列

        Args:
            row_index (int): 行索引

        Returns:
            int: 第一个可用列的列号
        """
        if not self.active_sheet:
            return 1

        for col in range(1, self.get_total_cols() + 1):
            cell_value = self.active_sheet.cell(row=row_index, column=col).value
            if cell_value is not None and str(cell_value).strip():
                return col

        return 1  # 如果没有找到，返回第一列

    def get_last_used_row(self) -> int:
        """
        获取最后一个使用的行

        Returns:
            int: 最后一个使用的行号
        """
        if not self.active_sheet:
            return 0

        # 从底部开始向上查找
        for row in range(self.get_total_rows(), 0, -1):
            for col in range(1, self.get_total_cols() + 1):
                cell_value = self.active_sheet.cell(row=row, column=col).value
                if cell_value is not None and str(cell_value).strip():
                    return row

        return 0

    def get_last_used_col(self) -> int:
        """
        获取最后一个使用的列

        Returns:
            int: 最后一个使用的列号
        """
        if not self.active_sheet:
            return 0

        # 从右侧开始向左查找
        for col in range(self.get_total_cols(), 0, -1):
            for row in range(1, self.get_total_rows() + 1):
                cell_value = self.active_sheet.cell(row=row, column=col).value
                if cell_value is not None and str(cell_value).strip():
                    return col

        return 0

    def clear_excel_content(self, start_row: int = 1, start_col: int = 1,
                            end_row: int = None, end_col: int = None):
        """
        清空Excel内容

        Args:
            start_row (int): 起始行
            start_col (int): 起始列
            end_row (int): 结束行
            end_col (int): 结束列
        """
        if not self.active_sheet:
            return

        # 设置默认结束位置
        if end_row is None:
            end_row = self.get_total_rows()
        if end_col is None:
            end_col = self.get_total_cols()

        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                self.active_sheet.cell(row=row, column=col).value = None

    def delete_excel_row(self, row_index: int):
        """
        删除Excel行

        Args:
            row_index (int): 要删除的行索引
        """
        if not self.active_sheet:
            return

        self.active_sheet.delete_rows(row_index)

    def delete_duplicate_rows(self, columns: List[int] = None):
        """
        删除重复行

        Args:
            columns (List[int]): 用于判断重复的列索引列表
        """
        if not self.active_sheet:
            return

        # 使用pandas处理重复行
        df = pd.read_excel(self.file_path)

        if columns:
            df.drop_duplicates(subset=[df.columns[i - 1] for i in columns], inplace=True)
        else:
            df.drop_duplicates(inplace=True)

        # 写回Excel
        df.to_excel(self.file_path, index=False)

    def delete_all_rows(self):
        """
        删除所有行
        """
        if not self.active_sheet:
            return

        self.active_sheet.delete_rows(1, self.get_total_rows())

    def delete_excel_column(self, col_index: int):
        """
        删除Excel列

        Args:
            col_index (int): 要删除的列索引
        """
        if not self.active_sheet:
            return

        self.active_sheet.delete_cols(col_index)

    def insert_empty_row(self, row_index: int):
        """
        插入空行

        Args:
            row_index (int): 插入位置的行索引
        """
        if not self.active_sheet:
            return

        self.active_sheet.insert_rows(row_index)

    def insert_empty_column(self, col_index: int):
        """
        插入空列

        Args:
            col_index (int): 插入位置的列索引
        """
        if not self.active_sheet:
            return

        self.active_sheet.insert_cols(col_index)

    def select_area(self, start_row: int, start_col: int, end_row: int, end_col: int):
        """
        选中区域

        Args:
            start_row (int): 起始行
            start_col (int): 起始列
            end_row (int): 结束行
            end_col (int): 结束列

        Returns:
            List[List]: 选中的区域内容
        """
        if not self.active_sheet:
            return []

        area = []
        for row in range(start_row, end_row + 1):
            row_data = []
            for col in range(start_col, end_col + 1):
                cell_value = self.active_sheet.cell(row=row, column=col).value
                row_data.append(cell_value)
            area.append(row_data)

        return area

    def get_selected_area(self, start_row: int, start_col: int, end_row: int, end_col: int):
        """
        获取选中区域的内容

        Args:
            start_row (int): 起始行
            start_col (int): 起始列
            end_row (int): 结束行
            end_col (int): 结束列

        Returns:
            List[List]: 选中区域的内容
        """
        return self.select_area(start_row, start_col, end_row, end_col)

    def set_column_hidden(self, col_index: int, hidden: bool = True):
        """
        设置列隐藏

        Args:
            col_index (int): 列索引
            hidden (bool): 是否隐藏
        """
        if not self.active_sheet:
            return

        self.active_sheet.column_dimensions[openpyxl.utils.get_column_letter(col_index)].hidden = hidden

    def set_row_hidden(self, row_index: int, hidden: bool = True):
        """
        设置行隐藏

        Args:
            row_index (int): 行索引
            hidden (bool): 是否隐藏
        """
        if not self.active_sheet:
            return

        self.active_sheet.row_dimensions[row_index].hidden = hidden

    def save_file(self):
        """保存文件"""
        if self.workbook:
            self.workbook.save(self.file_path)


# 使用示例
if __name__ == "__main__":
    # 创建Excel操作实例
    excel_op = ExcelOperations("example.xlsx")

    # 打开文件
    if excel_op.open_file():
        # 获取总行数
        total_rows = excel_op.get_total_rows()
        print(f"总行数: {total_rows}")

        # 获取总列数
        total_cols = excel_op.get_total_cols()
        print(f"总列数: {total_cols}")

        # 拷贝内容
        content = excel_op.copy_excel_content()
        print(f"拷贝了 {len(content)} 行数据")

        # 关闭文件
        excel_op.close_file()
