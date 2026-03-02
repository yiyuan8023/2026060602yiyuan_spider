import pandas as pd
import os
from datetime import datetime
from typing import List, Dict, Any
import openpyxl

from openpyxl.utils.dataframe import dataframe_to_rows
from extra.logger_ import logger


class SaveToExcel:
    """
    导出数据到excel
    """

    def __init__(
        self,
        items: List[Dict[str, Any]],
        filename: str = None,
        path: str = None,
        sheet_name: str = "Sheet1",
    ):
        """
        初始化Excel导出器
        Args:
            path (str, optional): 保存路径，默认为系统桌面
        """
        self.path = self._get_path(path)
        self.items = self._get_items(items)
        self.filename = self._get_filename(filename)
        self.sheet_name = sheet_name  # 默认为"Sheet1"
        self.file_path = os.path.join(self.path, self.filename)  # 文件路径
        self.df = pd.DataFrame(self.items)

    @staticmethod
    def _get_path(path: str = None) -> str:
        if path is None:
            # 获取系统桌面路径
            return os.path.join(os.path.expanduser("~"), "Desktop")
        else:
            # 确保路径存在
            if not os.path.exists(path):
                os.makedirs(path, exist_ok=True)
            return path

    @staticmethod
    def _get_items(items: List[Dict[str, Any]]):
        # 数据验证
        if not isinstance(items, list):
            raise ValueError("数据必须是列表类型")

        if not items:
            raise ValueError("数据不能为空")
        # 返回验证通过的数据
        return items

    @staticmethod
    def _get_filename(filename: str = None) -> str:

        # 生成默认文件名
        # 文件名，默认为"文件导出+时间戳"
        if filename is None:
            # 文件名，默认为"文件导出+时间戳"
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"文件导出_{timestamp}.xlsx"
        elif not filename.endswith((".xlsx", ".xls")):
            filename = f"{filename}.xlsx"
        return filename

    def export_to_excel(
        self,
    ) -> str:
        """
        将列表字典导出到Excel文件
        """
        try:
            # 检查文件是否存在
            if os.path.exists(self.file_path):
                # 如果文件存在，追加数据到第一个可用行
                self._append_to_existing_file()
            else:
                # 如果文件不存在，创建新文件
                self.df.to_excel(
                    self.file_path, sheet_name=self.sheet_name, index=False
                )

            logger.info(f"✅ 数据已成功导出到: {self.file_path}")
            logger.info(f"📊 导出记录数: {len(self.items)}")
            logger.info(f"📋 导出列数: {len(self.df.columns)}")

            return self.file_path

        except Exception as exc:
            error_msg = f"导出Excel文件时出错: {str(exc)}"
            logger.info(f"❌ {error_msg}")
            raise Exception(error_msg)

    def _append_to_existing_file(self):
        """
        追加数据到已存在的Excel文件
        """
        try:
            # 加载现有工作簿
            workbook = openpyxl.load_workbook(self.file_path)

            # 检查工作表是否存在
            if self.sheet_name in workbook.sheetnames:
                worksheet = workbook[self.sheet_name]
            else:
                worksheet = workbook.create_sheet(self.sheet_name)

            # 找到第一个可用行（跳过标题行）
            first_empty_row = worksheet.max_row + 1
            if first_empty_row == 1:
                # 如果工作表为空，先写入标题
                for col_idx, column_name in enumerate(self.df.columns, 1):
                    worksheet.cell(row=1, column=col_idx, value=column_name)
                first_empty_row = 2

            # 追加数据（不包含标题）
            for row_data in dataframe_to_rows(self.df, index=False, header=False):
                for col_idx, value in enumerate(row_data, 1):
                    worksheet.cell(row=first_empty_row, column=col_idx, value=value)
                first_empty_row += 1

            # 保存工作簿
            workbook.save(self.file_path)

        except Exception as exc:
            # 如果追加失败，直接覆盖文件
            logger.warning(f"追加数据失败，将创建新文件: {str(exc)}")
            self.df.to_excel(self.file_path, sheet_name=self.sheet_name, index=False)


# 使用示例
if __name__ == "__main__":
    # 创建导出器实例

    # 示例数据
    sample_data = [
        {"姓名": "张三", "年龄": 25, "部门": "技术部", "薪资": 8000},
        {"姓名": "李四", "年龄": 30, "部门": "销售部", "薪资": 9000},
    ]
    exporter = SaveToExcel(sample_data, "员工信息表")

    # 示例1: 基本导出
    try:
        file_path = exporter.export_to_excel()
        print(f"文件已保存到: {file_path}")
    except Exception as e:
        print(f"导出失败: {e}")
